import openpyxl as op
import os

os.system('cls')
if os.path.exists("favorite_people.xlsx"):
    wbk = op.load_workbook("favorite_people.xlsx")
    sheet = wbk.active
else:
    wbk = op.Workbook()
    sheet = wbk.active
    sheet["A1"] = "ID"
    sheet["B1"] = "First Name"
    sheet["C1"] = "Last Name"
    sheet["D1"] = "Birth Year"
    sheet["E1"] = "Age"
    sheet["F1"] = "Next Append"
    

cell_value = sheet['G1'].value
if cell_value is not None:
    num_people = int(cell_value)
else:
    num_people = 0

print("\nHello user please Enter your favorite person\ntheir First and Last name, lastly their Birthyear!\n")

while True:
    print(f"\nPerson {num_people + 1}: ")
    ID = num_people + 1
    FirstName = input("First Name: ")
    LastName = input("Last Name: ")
    try:
        BirthYear = int(input("Birth Year: "))
    except ValueError:
        print("Invalid input. Please enter a valid year.")
        continue

    Age = 2026 - BirthYear
    append = [ID, FirstName, LastName, BirthYear, Age]
    sheet.append(append)

    Continue = input("Do you want to add another person? (yes/no)").lower()
    if Continue == "yes":
        num_people += 1
        continue
    else:
        sheet[f"G1"] = num_people + 1
        print("\nFinished adding people.\n")
        break

print("=== Your Favorite People ===\n")

for rows in sheet.iter_rows(max_col=5,values_only=True):
    print(rows)

print("\n")

input("Enter to save and exit.")

wbk.save("favorite_people.xlsx")